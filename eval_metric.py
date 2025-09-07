# -*- coding: utf-8 -*-
"""
高光谱图像（HSI）重建结果评估脚本
功能：计算不同噪声场景下的图像质量指标（PSNR、SSIM、ERGAS、SAM），并将结果保存为Excel表格

文章：
Shuang Xu, Zixiang Zhao, Haowen Bai, Chang Yu, Jiangjun Peng, Xiangyong Cao, Deyu Meng, 
Hipandas: Hyperspectral Image Joint Denoising and Super-Resolution by Image Fusion with the Panchromatic Image. 
ICCV, 2025.
"""

# 导入所需库
import scipy.io as sio  # 用于读取和保存.mat格式文件
from utils import *     # 导入自定义工具函数（如HSI_metrics）
import numpy as np      # 用于数值计算
import openpyxl         # 用于操作Excel文件

# 定义评估的方法名称（对应结果文件夹）
method = 'UHipandas'

# 定义噪声场景列表（高斯噪声、混合噪声等）
noise_case = ['g10', 'g30',  'gni', 'mix15', 'mix35', 'mix55']
# 初始化字典存储各噪声场景的评估指标
metric = dict([(k,[]) for k in noise_case])

# 遍历所有9x9=81幅测试图像
for i in range(9):
    for j in range(9):
        filename = 'Dongying_%d_%d.mat'%(i,j)  # 图像文件名
        print(filename)  # 打印当前处理的文件名
        img_id = '%d_%d'%(i,j)  # 图像ID（用于表格标识）
        
        # 遍历每个噪声场景
        for nc in noise_case:
            # 构建真实图像（GT）路径
            gt_path = 'data/%s/%s'%(nc,filename) 
            # 读取GT数据并归一化到[0,1]范围（原始数据为10位量化，最大值1023）
            I_GT = sio.loadmat(gt_path)['I_GT'].astype('float')/1023
            
            # 构建重建结果路径
            recon_path = 'result/%s/%s/%s'%(method,nc,filename) 
            # 读取重建结果并归一化
            Output = sio.loadmat(recon_path)['output'].astype('float')/1023
            
            # 计算评估指标（PSNR、SSIM、ERGAS、SAM）
            m = HSI_metrics(I_GT, Output)
            # 在指标列表前插入图像ID
            m.insert(0,img_id)
            # 将结果存入对应噪声场景的列表
            metric[nc].append(m)

# 创建Excel工作簿
data = openpyxl.Workbook()
# 删除默认工作表
del data['Sheet']
# 存储各噪声场景的平均指标
avg_metric = []

# 为每个噪声场景创建工作表并写入数据
for nc in list(metric.keys()):
    table = data.create_sheet(nc)  # 创建以噪声场景命名的工作表
    # 写入表头
    table.append(['ID', 'PSNR', 'SSIM', 'ERGAS', 'SAM'])
    # 计算当前噪声场景下所有图像的平均指标并存储
    avg_metric.append(np.array(metric[nc])[:,1:].astype('float').mean(0).tolist())
    # 写入每个图像的指标数据
    for row in metric[nc]:
        table.append(row)

# 创建汇总工作表，存储各噪声场景的平均指标
table = data.create_sheet('summary')
table.append(['Case', 'PSNR', 'SSIM', 'ERGAS', 'SAM'])
# 写入各噪声场景的平均指标
for nc, row in zip(noise_case,avg_metric):
    row.insert(0,nc)  # 在平均指标前插入噪声场景名称
    table.append(row)

# 保存Excel文件到结果目录
data.save('result/%s.xlsx'%(method))